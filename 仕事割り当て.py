import streamlit as st
from pulp import *
from openpyxl import *
import pandas as pd
import time
from io import BytesIO
import tempfile

#Excelシートの読み込み
st.session_state.book = load_workbook('卒業研究_シフトデータ.xlsx')
sh1 = st.session_state.book['入力データ']
sh2 = st.session_state.book['シフト表']

st.title('ライブ裏方仕事の割り当て')
st.header('▼ 初期情報の入力')

@st.cache_resource
def input_list_member():
    st.session_state.lst_member = []
    return st.session_state.lst_member

@st.cache_resource
def input_dict_turn_member():
    dict_turn_member = {}
    return dict_turn_member

#コールバック関数
def clear_member():
    input_list_member.clear()
    st.session_state.lst_member = input_list_member()

#メンバーの名前入力
st.session_state.lst_member = input_list_member()
dict_turn_member = input_dict_turn_member()
with st.form('input_name', clear_on_submit=True):
    st.write('メンバーの名前を入力してください')
    member_name = st.text_input('↓↓↓')
    submitted = st.form_submit_button('メンバー追加')
    if submitted:
        st.session_state.lst_member.append(member_name)
        dict_turn_member[member_name] = 0
    st.table(st.session_state.lst_member)
member = len(st.session_state.lst_member)
btn = st.button('clear member', on_click=clear_member)

#バンド数入力
slot = st.number_input('出演バンド数を決めてください', min_value=1, max_value=100)
st.write('→ 出演バンドは', slot, 'バンドです。')
    
#定数用のデータの作成
st.session_state.I = [i+1 for i in range(member)]
st.session_state.J = [i+1 for i in range(4)]
st.session_state.T = [i+1 for i in range(slot)]

if int(member) > 0:
        
#コールバック関数
    def update_index():
        st.session_state['num'] += 1

#コールバック関数
    def clear_turn():
        input_dict_turn_member.clear()
        dict_turn_member = input_dict_turn_member()
        st.session_state['num']= 0

#出演バンドの入力
    if len(dict_turn_member) > 0:
        with st.form('input_turn', clear_on_submit=True):
            if 'num' not in st.session_state:
                st.session_state['num'] = 0
            if st.session_state['num'] < int(member):
                st.write(st.session_state.lst_member[st.session_state['num']], 'が何バンド目に出演するか入力してください')
                st.write('出演しない人は0を入力してください')
            if st.session_state['num'] == int(member):
                st.write('入力完了')
                lst_turn = list(dict_turn_member.values())
            member_turn = st.number_input('↓↓↓', min_value=0, max_value=slot)
            submitted = st.form_submit_button('出演情報追加', on_click=update_index)
            if submitted:
                dict_turn_member[st.session_state.lst_member[st.session_state['num']-1]] = member_turn
            st.table(dict_turn_member)
        btn = st.button('clear turn', on_click=clear_turn)

    #Excel書き込み
    if len(dict_turn_member) > 0:
            for i in range(0, int(slot)):
                sh1.cell(row=1, column=2+i).value = i+1
                sh2.cell(row=1, column=2+i).value = i+1
            for i in range(0, len(st.session_state.lst_member)):
                sh1.cell(row=2+i, column=1).value = st.session_state.lst_member[i]
                for t in range(0, int(slot)):
                    sh1.cell(row=2+i, column=2+t).value = 1
            for i in range(0, len(lst_turn)):
                if lst_turn[i] > 0:
                    sh1.cell(row=2+i, column=1+lst_turn[i]).value = 0
            for i in range(0, len(lst_turn)):
                if lst_turn[i] < int(slot) and lst_turn[i] > 0:
                    sh1.cell(row=2+i, column=2+lst_turn[i]).value = 2
            for i in range(0, len(lst_turn)):
                if lst_turn[i] > 1:
                    sh1.cell(row=2+i, column=lst_turn[i]).value = 3
            st.session_state.book.save('卒業研究_シフトデータ_1.xlsx')
            #df = pd.read_excel('卒業研究_シフトデータ_1.xlsx', sheet_name='入力データ', index_col=0)
            #st.dataframe(df)
        
#定数用のデータの作成        
            a = {}    
            for i in st.session_state.I:
                for t in st.session_state.T:
                    a[i, t] = sh1.cell(row=1+i, column=1+t).value      
    
#空問題の作成
            model = LpProblem('Shift', sense=LpMinimize)

#決定変数の作成
            x = {}
            for i in st.session_state.I:
                for j in st.session_state.J:
                    for t in st.session_state.T:
                        x[i, j, t] = LpVariable(f'x{i},{j},{t}', cat=LpBinary)
            v = {}
            for i in st.session_state.I:
                v[i] = LpVariable(f'v{i}', lowBound=0)
            w = {}
            for i in st.session_state.I:
                for t in st.session_state.T:
                    w[i, t] = LpVariable(f'w{i}, {t}', cat=LpBinary)
            z = {}
            for i in st.session_state.I:
                for t in st.session_state.T:
                    z[i, t] = LpVariable(f'z{i}, {t}', cat=LpBinary)

#制約条件の追加
#条件①
            for j in st.session_state.J:
                for t in st.session_state.T:
                    model += lpSum(x[i, j, t] for i in st.session_state.I) == 1
#条件②
            for i in st.session_state.I:
                for t in st.session_state.T:
                    model += lpSum(x[i, j, t] for j in st.session_state.J) <= 1
#条件③
            for i in st.session_state.I:
                for t in st.session_state.T:
                    if a[i, t] == 0:
                        model += lpSum(x[i, j, t] for j in st.session_state.J) == 0
#条件④
            for i in st.session_state.I:
                for t in st.session_state.T:
                    if a[i, t] == 3:
                        model += lpSum(x[i, j, t] for j in st.session_state.J) == 0
#条件⑤
            for i in st.session_state.I:
                for t in st.session_state.T:
                    if a[i, t] == 2:
                        model += lpSum(x[i, j, t] for j in st.session_state.J) == 0 + w[i, t]
#条件⑥
            s = (len(st.session_state.J) * len(st.session_state.T)) / len(st.session_state.I)
            for i in st.session_state.I:
                model += lpSum(x[i, j, t] for j in st.session_state.J for t in st.session_state.T) >= s - v[i]
                model += lpSum(x[i, j, t] for j in st.session_state.J for t in st.session_state.T) <= s + v[i]
#条件⑦
            for i in st.session_state.I:
                for t in range(len(st.session_state.T)-1):
                    model += lpSum(x[i, j, t+1] for j in st.session_state.J) + lpSum(x[i, j, t+2] for j in st.session_state.J) <= 1 + z[i, t+1]
        
#目的関数の設定
            model += lpSum(v[i] for i in st.session_state.I) + lpSum(z[i, t] for i in st.session_state.I for t in st.session_state.T) + lpSum(w[i, t] for i in st.session_state.I for t in st.session_state.T)

#最適化の実行
            model.solve()

            st.header('▼ 裏方仕事の割り当て表示')
            st.write('最適値を求めて裏方仕事の割り当てを決定します。')

#実行結果の表示
            if st.button('仕事割り当て'):
                status = st.empty()
                bar = st.progress(0)
                for i in range(100):
                    status.text(f'Status {i+1}')
                    bar.progress(i + 1)
                    time.sleep(0.05)

                if LpStatus[model.status] == 'Optimal':
                    I_Name = dict(zip(st.session_state.I, st.session_state.lst_member))
                    for i in st.session_state.I:
                        for j in st.session_state.J:
                            for t in st.session_state.T:
                                if value(x[i, j, t]) > 0.01:
                                    sh2.cell(row=1+j, column=1+t).value = I_Name[i]    #シフト表に名前を出力
                    st.session_state.book.save('卒業研究_シフトデータ_2.xlsx')
                    df = pd.read_excel('卒業研究_シフトデータ_2.xlsx', sheet_name='シフト表', index_col=0)
                    st.dataframe(df)

                    workbook = load_workbook('卒業研究_シフトデータ_2.xlsx')
                    with tempfile.NamedTemporaryFile(delete=False) as tmp:
                        workbook.save(tmp.name)
                        data = BytesIO(tmp.read())
                    st.download_button("データをダウンロード",data=data,mime='xlsx',file_name="割り当てデータ.xlsx")
                else:
                    st.write('裏方仕事の割り当てを表示できませんでした')