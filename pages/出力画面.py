import streamlit as st
from pulp import *
from openpyxl import *
import pandas as pd
import time

st.title('ライブ裏方シフトのシフト表')

#シフト情報の読み込み
st.session_state.book = load_workbook('卒業研究_シフトデータ_1.xlsx')
sh1 = st.session_state.book['入力データ']
sh2 = st.session_state.book['シフト表']

uploaded_file = st.file_uploader("CSVファイルをアップロードしてください", accept_multiple_files=True)
#if uploaded_file is not None:
    #data = pd.read_csv(uploaded_file)
    #st.dataframe(data)
for uploaded_file in uploaded_files:
    st.dataframe(uploaded_file.name)



#定数用のデータの作成
a = {}    
for i in st.session_state.I:
    for t in st.session_state.T:
        a[i, t] = sh1.cell(row=1+i, column=1+t).value      
    
#空問題の作成
model = LpProblem('Shift', sense=LpMinimize)

#決定変数の作成
x = {}
for i in st.session_state.I:
    for j in st.session_state.J:
        for t in st.session_state.T:
            x[i, j, t] = LpVariable(f'x{i},{j},{t}', cat=LpBinary)
v = {}
for i in st.session_state.I:
    v[i] = LpVariable(f'v{i}', lowBound=0)
w = {}
for i in st.session_state.I:
    for t in st.session_state.T:
        w[i, t] = LpVariable(f'w{i}, {t}', cat=LpBinary)
z = {}
for i in st.session_state.I:
    for t in st.session_state.T:
        z[i, t] = LpVariable(f'z{i}, {t}', cat=LpBinary)

#制約条件の追加

#条件①
for j in st.session_state.J:
    for t in st.session_state.T:
        model += lpSum(x[i, j, t] for i in st.session_state.I) == 1
#条件②
for i in st.session_state.I:
    for t in st.session_state.T:
        model += lpSum(x[i, j, t] for j in st.session_state.J) <= 1
#条件③
for i in st.session_state.I:
    for t in st.session_state.T:
        if a[i, t] == 0:
            model += lpSum(x[i, j, t] for j in st.session_state.J) == 0
#条件④
for i in st.session_state.I:
    for t in st.session_state.T:
        if a[i, t] == 3:
            model += lpSum(x[i, j, t] for j in st.session_state.J) == 0
#条件⑤
s = (len(st.session_state.J) * len(st.session_state.T)) / len(st.session_state.I)
for i in st.session_state.I:
    model += lpSum(x[i, j, t] for j in st.session_state.J for t in st.session_state.T) >= s - v[i]
    model += lpSum(x[i, j, t] for j in st.session_state.J for t in st.session_state.T) <= s + v[i]
#条件⑥
for i in st.session_state.I:
    for t in range(len(st.session_state.T)-1):
        model += lpSum(x[i, j, t+1] for j in st.session_state.J) + lpSum(x[i, j, t+2] for j in st.session_state.J) <= 1 + z[i, t+1]
#条件⑦
for i in st.session_state.I:
    for t in st.session_state.T:
        if a[i, t] == 2:
            model += lpSum(x[i, j, t] for j in st.session_state.J) == 0 + w[i, t]
        
#目的関数の設定
model += lpSum(v[i] for i in st.session_state.I) + lpSum(z[i, t] for i in st.session_state.I for t in st.session_state.T) + lpSum(w[i, t] for i in st.session_state.I for t in st.session_state.T)

#最適化の実行
model.solve()

st.header('▼ シフトを作成する')
st.write('最適値を求めてシフトを作成します。')

#実行結果の表示
if st.button('シフト作成'):
    status = st.empty()
    bar = st.progress(0)
    for i in range(100):
        status.text(f'Status {i+1}')
        bar.progress(i + 1)
        time.sleep(0.05)
        
    if LpStatus[model.status] == 'Optimal':
        I_Name = dict(zip(st.session_state.I, st.session_state.lst_member))
        for i in st.session_state.I:
            for j in st.session_state.J:
                for t in st.session_state.T:
                    if value(x[i, j, t]) > 0.01:
                        sh2.cell(row=1+j, column=1+t).value = I_Name[i]    #シフト表に名前を出力
        st.session_state.book.save('卒業研究_シフトデータ_2.xlsx')
        df = pd.read_excel('卒業研究_シフトデータ_2.xlsx', sheet_name='シフト表', index_col=0)
        st.dataframe(df)
        df_shift=pd.DataFrame(data=df)
        st.download_button(label='シフトをダウンロード', data=df_shift.to_csv().encode('utf-8-sig'), file_name='シフト表.csv')
    else:
        st.write('シフトが作成できませんでした')